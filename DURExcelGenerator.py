#! DURExcelGenerator.py - A simple python script to combine multiple excels output from MHSG
#  into a single excel file that also counts the number of times a medication was dispensed
#  and how much in total of that medication was dispensed.

import openpyxl
from openpyxl.styles import Font, Alignment
import glob
from datetime import date


# A variable for assigning the year to the output file
dur_year = date.today().year - 1

# Using glob to make a list of all the excel files in the current directory.
files = glob.glob('*.xlsx')

# Making an array to accept an array for each line, so each occurrence, from each excel document.
occurrences = []

def get_item_id(occurrence):
    """Going to return the item id from each occurrence to be used to sort the occurrences."""
    item_id = occurrence[0]
    return item_id

# Setting up a loop to iterate over every excel file identified above.
for file in files:
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # Setting up a loop to iterate over every row in the excel document; will grab the item id (B), description (C),
    # dispenses (G), and quantity(H). Starting at 2 to skip the header information and adding + 1 to max_row
    # in order to get the last item of each sheet.
    for row in range(2, sheet.max_row + 1):
        item_id = sheet[f'B{row}'].value
        description = sheet[f'C{row}'].value
        dispenses = sheet[f'G{row}'].value
        quantity = sheet[f'H{row}'].value
        occurrence = [item_id, description, dispenses, quantity]
        occurrences.append(occurrence)

    # Giving some feedback while running program and closing the workbook before starting the next one.
    print(f'Finished processing {file}')
    wb.close()

# Sorting the data before we can count the times dispensed and quantity dispensed for each item id.
sorted_occurences = sorted(occurrences, key=get_item_id)

# Setting up variables to start identifying multiple occurrences and adding into one output value.
item_id = sorted_occurences[0][0]
description = sorted_occurences[0][1]
dispenses = 0
quantity = 0

# Open new workbook to write to and set active sheet.
wb = openpyxl.Workbook()
sheet = wb.active

# Setting up font variables for the title and headers
fontObjTitle = Font(name='Times New Roman', size=16, bold=True)
fontObjHeader = Font(name='Times New Roman', size=14)

# Formatting the columns
sheet.merge_cells('A1:D1')
sheet['A1'].font = fontObjTitle
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['A2'].font = fontObjHeader
sheet['B2'].font = fontObjHeader
sheet['C2'].font = fontObjHeader
sheet['D2'].font = fontObjHeader

# Writing the title and header information
sheet['A1'] = f'Drug Utilization Review: {dur_year}'
sheet['A2'] = 'Item ID'
sheet['B2'] = 'Description'
sheet['C2'] = 'Dispenses'
sheet['D2'] = 'Quantity Dispensed'


row_to_write_to = 3

# Loop through all occurrences in sorted occurrence. While item_id is the same add to the current dispenses
# and quantity variables.
for occurrence in sorted_occurences:
    if occurrence[0] == item_id:
        dispenses += occurrence[2]
        quantity += occurrence[3]
    else:
        # Write values to new workbook
        sheet[f'A{row_to_write_to}'] = item_id
        sheet[f'B{row_to_write_to}'] = description
        sheet[f'C{row_to_write_to}'] = dispenses
        sheet[f'D{row_to_write_to}'] = quantity

        # Increment the row to write to next
        row_to_write_to += 1

        # Set new variables with the next values
        if occurrence[0] is not None:
            item_id = occurrence[0]
            description = occurrence[1]
            dispenses = occurrence[2]
            quantity = occurrence[3]
        else:
            break

wb.save(f'DUR{dur_year}.xlsx')
wb.close()